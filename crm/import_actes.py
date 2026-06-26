"""Import / export du referentiel d'actes via un fichier Excel.

Alimente la table `actes` (referentiel tarife, cf. crm/db.py v9) a partir d'un
classeur .xlsx contenant (id, libelle, prix/montant, code, categorie, actif) —
id / code / categorie / actif facultatifs. La categorie classe les actes (ex.
« Prothese », « Chirurgie ») ; cf. crm/db.py v13 (colonne `actes.categorie`).

Flux aller-retour (« export -> j'edite -> reimport ») : on exporte le referentiel,
on edite des lignes ou on en ajoute, on reimporte. Les lignes existantes sont mises
a jour, les nouvelles sont creees.

Idempotent : relancer le meme fichier MET A JOUR les actes existants au lieu de
creer des doublons. Cle de rapprochement (pour reconnaitre un acte deja present),
par ordre de priorite :

  1. l'ID (colonne « ID ») s'il est renseigne — cle stable de l'export, seule a
     survivre a une edition du libelle ET du code. Un ID renseigne mais introuvable
     en base est signale (la ligne est ignoree, pas de creation surprise) ;
  2. sinon le CODE s'il est renseigne (cle stable, ex. nomenclature) ;
  3. sinon le LIBELLE (insensible aux accents et a la casse, comme dans l'app).

La colonne ID est ecrite a l'export et NE DOIT PAS etre modifiee a la main : c'est
elle qui fait le lien avec l'acte existant. Laisser la cellule ID vide sur une
nouvelle ligne -> l'acte est cree.

Format attendu du classeur (premiere feuille par defaut) :

  - AVEC une ligne d'entete : les colonnes sont reconnues par leur nom
    (« id », « libelle / acte / designation », « prix / montant / tarif »,
    « code / reference », « categorie / famille / groupe », « actif / etat »),
    quel que soit leur ordre. Les colonnes id / code / categorie / actif sont
    facultatives.
  - SANS entete : colonnes par position -> 1 = libelle, 2 = prix, 3 = code,
    4 = categorie (pas d'ID en mode positionnel : un fichier saisi a la main sans
    entete decrit toujours des actes a creer/rapprocher par code ou libelle).

Les montants acceptent le format francais (« 1 800,00 », « 1800.00 », « 120 € »...).

Usage :
    python -m crm.import_actes actes.xlsx
    python -m crm.import_actes actes.xlsx --feuille "Tarifs"
    python -m crm.import_actes actes.xlsx --dry-run     # simulation, n'ecrit rien
    python -m crm.import_actes --export referentiel.xlsx           # export du referentiel
    python -m crm.import_actes --export referentiel.xlsx --inclure-inactifs
    python -m crm.import_actes --modele modele_actes.xlsx   # genere un fichier exemple

Une sauvegarde de la base est prise avant tout ecriture (cf. crm/backup.py).
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Optional

from . import repo
from .backup import backup_db
from .db import connect, default_db_path
from .repo import slugify

# Interface terminal coloree partagee (src/ui.py), comme crm/reset.py.
try:
    from src import ui
except Exception:  # noqa: BLE001 - fallback si src indisponible
    ui = None  # type: ignore[assignment]


# Mots-cles d'entete reconnus (compares au slug de la cellule d'entete).
# ATTENTION : ne pas faire chevaucher _ID_KEYS et _CODE_KEYS (« reference » est un
# code, pas l'ID interne) — l'ID est la cle technique de l'export.
_ID_KEYS = {
    "id", "ids", "identifiant", "identifiants", "ident", "id interne",
    "identifiant interne", "ref interne", "reference interne",
}
_LIBELLE_KEYS = {
    "libelle", "libelles", "label", "acte", "actes", "designation",
    "intitule", "nom", "prestation", "description",
}
_PRIX_KEYS = {
    "prix", "montant", "montants", "tarif", "tarifs", "cout", "couts",
    "price", "amount", "honoraire", "honoraires",
}
_CODE_KEYS = {
    "code", "codes", "reference", "ref", "nomenclature", "ngap", "ccam",
}
_CATEGORIE_KEYS = {
    "categorie", "categories", "category", "famille", "familles", "groupe",
    "groupes", "rubrique", "rubriques", "classe", "classes", "specialite",
    "specialites",
}
_ACTIF_KEYS = {
    "actif", "actifs", "active", "etat", "etats", "statut", "statuts", "status",
}

# Valeurs textuelles reconnues pour la colonne « Actif » (slug, sans accents).
_ACTIF_VRAI = {"oui", "o", "yes", "y", "vrai", "true", "1", "actif", "active", "x"}
_ACTIF_FAUX = {"non", "n", "no", "faux", "false", "0", "inactif", "inactive", "desactive"}


@dataclass
class ImportSummary:
    created: int = 0
    updated: int = 0
    skipped: int = 0          # lignes vides / ignorees (libelle manquant, ID inconnu)
    errors: list[str] = None  # type: ignore[assignment]

    def __post_init__(self) -> None:
        if self.errors is None:
            self.errors = []


# --- Lecture / normalisation des cellules --------------------------------------

def _parse_montant(value: object) -> Optional[float]:
    """Convertit une cellule en nombre, en tolerant le format francais.

    Renvoie None si la cellule est vide. Leve ValueError si le contenu n'est pas
    interpretable comme un montant.
    """
    if value is None:
        return None
    if isinstance(value, bool):  # bool est un int en Python : on refuse
        raise ValueError("valeur booleenne")
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    # Retire espaces (y compris insecables) et tout sauf chiffres , . -
    s = s.replace(" ", "").replace("\xa0", "").replace(" ", "")
    s = re.sub(r"[^0-9,.\-]", "", s)
    if not s or s in {"-", ".", ","}:
        raise ValueError(f"montant illisible : {value!r}")
    # Si , et . coexistent, le dernier separateur est le decimal.
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError as exc:
        raise ValueError(f"montant illisible : {value!r}") from exc


def _clean_libelle(value: object) -> str:
    return "" if value is None else str(value).strip()


def _clean_categorie(value: object) -> Optional[str]:
    """Categorie en texte (None si cellule vide). Normalisation finale (trim/None)
    cote repo ; ici on extrait juste une chaine exploitable."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))  # un code/numero de categorie « 1.0 » -> « 1 »
    s = str(value).strip()
    return s or None


def _parse_code(value: object) -> Optional[str]:
    """Code en texte. Un code numerique entier (123.0) devient « 123 »."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    s = str(value).strip()
    return s or None


def _parse_id(value: object) -> Optional[int]:
    """ID interne en entier (None si cellule vide). Leve ValueError si non entier."""
    if value is None:
        return None
    if isinstance(value, bool):
        raise ValueError("ID invalide")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        raise ValueError(f"ID invalide : {value!r}")
    s = str(value).strip()
    if not s:
        return None
    if not re.fullmatch(r"\d+", s):
        raise ValueError(f"ID invalide : {value!r}")
    return int(s)


def _parse_actif(value: object) -> Optional[bool]:
    """Etat actif/inactif (None si cellule vide ou non reconnue : on ne touche pas)."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    s = slugify(str(value))
    if s in _ACTIF_VRAI:
        return True
    if s in _ACTIF_FAUX:
        return False
    return None


def _detect_columns(header: tuple) -> Optional[dict[str, Optional[int]]]:
    """Detecte les colonnes a partir d'une ligne d'entete.

    Renvoie {'id': .., 'libelle': .., 'prix': .., 'code': .., 'categorie': ..,
    'actif': ..} (chaque valeur = index ou None) si l'entete contient au moins une
    colonne libelle OU prix reconnue, sinon None (pas d'entete -> mode positionnel).
    """
    cols: dict[str, Optional[int]] = {
        "id": None, "libelle": None, "prix": None, "code": None,
        "categorie": None, "actif": None,
    }
    for i, cell in enumerate(header):
        slug = slugify(str(cell)) if cell is not None else ""
        if not slug:
            continue
        if cols["id"] is None and slug in _ID_KEYS:
            cols["id"] = i
        elif cols["libelle"] is None and slug in _LIBELLE_KEYS:
            cols["libelle"] = i
        elif cols["prix"] is None and slug in _PRIX_KEYS:
            cols["prix"] = i
        elif cols["code"] is None and slug in _CODE_KEYS:
            cols["code"] = i
        elif cols["categorie"] is None and slug in _CATEGORIE_KEYS:
            cols["categorie"] = i
        elif cols["actif"] is None and slug in _ACTIF_KEYS:
            cols["actif"] = i
    if cols["libelle"] is None and cols["prix"] is None:
        return None  # aucune entete reconnue
    # Valeurs par defaut pour une colonne manquante (rare : entete partielle).
    if cols["libelle"] is None:
        cols["libelle"] = 0
    if cols["prix"] is None:
        cols["prix"] = 1
    return cols


# Une ligne lue : (no_ligne, valeur_id, libelle, valeur_prix, valeur_code,
# valeur_categorie, valeur_actif).
_Row = tuple[int, object, str, object, object, object, object]


def _read_rows(path: Path, feuille: Optional[str]) -> list[_Row]:
    """Lit le classeur et renvoie la liste des lignes brutes a importer.

    Le no_ligne (1-based, tel qu'affiche dans Excel) sert aux messages d'erreur.
    """
    try:
        import openpyxl  # import tardif : dependance optionnelle
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "Le module 'openpyxl' est requis pour lire un .xlsx "
            "(pip install openpyxl)."
        ) from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if feuille:
            if feuille not in wb.sheetnames:
                raise RuntimeError(
                    f"Feuille « {feuille} » introuvable. "
                    f"Feuilles disponibles : {', '.join(wb.sheetnames)}"
                )
            ws = wb[feuille]
        else:
            ws = wb.active

        all_rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not all_rows:
        return []

    cols = _detect_columns(all_rows[0])
    if cols is not None:
        data_rows = all_rows[1:]
        start_line = 2  # la ligne 1 est l'entete
        i_id, i_lib, i_prix = cols["id"], cols["libelle"], cols["prix"]
        i_code, i_cat, i_actif = cols["code"], cols["categorie"], cols["actif"]
    else:
        data_rows = all_rows
        start_line = 1
        i_id = None  # pas d'ID en mode positionnel
        i_lib, i_prix, i_code, i_cat, i_actif = 0, 1, 2, 3, None  # positionnel

    out: list[_Row] = []
    for offset, row in enumerate(data_rows):
        line_no = start_line + offset

        def _at(idx: Optional[int]):
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        libelle = _clean_libelle(_at(i_lib))
        out.append((line_no, _at(i_id), libelle, _at(i_prix), _at(i_code),
                    _at(i_cat), _at(i_actif)))
    return out


# --- Rapprochement avec un acte existant ---------------------------------------

def _find_existing(
    conn, acte_id: Optional[int], code: Optional[str], libelle: str
) -> Optional[repo.Acte]:
    """Acte deja present correspondant a la ligne, ou None.

    Priorite a l'ID (cle stable de l'export), puis au code (actifs ET inactifs), a
    defaut le libelle parmi les actes actifs (regle de doublon de l'application).
    L'ID est traite par l'appelant (cas « ID fourni mais introuvable »).
    """
    if acte_id is not None:
        return repo.get_acte(conn, acte_id)
    if code:
        row = conn.execute(
            "SELECT id FROM actes WHERE code = ? ORDER BY id LIMIT 1", (code,)
        ).fetchone()
        if row:
            return repo.get_acte(conn, row["id"])
    return repo.find_acte_by_libelle(conn, libelle)


# --- Import --------------------------------------------------------------------

def _import_rows(conn, rows: list[_Row], dry_run: bool) -> ImportSummary:
    """Applique les lignes lues sur la connexion fournie (cle de rapprochement :
    ID -> code -> libelle)."""
    summary = ImportSummary()
    for line_no, raw_id, libelle, raw_prix, raw_code, raw_cat, raw_actif in rows:
        if (not libelle and raw_prix is None and raw_code is None
                and raw_cat is None and raw_id is None):
            summary.skipped += 1  # ligne entierement vide
            continue

        try:
            acte_id = _parse_id(raw_id)
        except ValueError as exc:
            summary.errors.append(f"Ligne {line_no} ({libelle or '?'}) : {exc}, ignoree.")
            summary.skipped += 1
            continue

        if not libelle:
            summary.errors.append(f"Ligne {line_no} : libelle manquant, ignoree.")
            summary.skipped += 1
            continue
        try:
            prix = _parse_montant(raw_prix)
        except ValueError as exc:
            summary.errors.append(f"Ligne {line_no} ({libelle}) : {exc}, ignoree.")
            summary.skipped += 1
            continue
        if prix is None:
            prix = 0.0  # libelle sans prix -> acte a 0 (ex. controle)
        if prix < 0:
            summary.errors.append(
                f"Ligne {line_no} ({libelle}) : prix negatif, ignoree."
            )
            summary.skipped += 1
            continue

        code = _parse_code(raw_code)
        categorie = _clean_categorie(raw_cat)
        actif = _parse_actif(raw_actif)
        existing = _find_existing(conn, acte_id, code, libelle)

        # ID fourni mais introuvable : on NE cree PAS (l'utilisateur a sans doute
        # modifie / mal saisi un ID issu de l'export, ou l'acte a ete supprime).
        if acte_id is not None and existing is None:
            summary.errors.append(
                f"Ligne {line_no} ({libelle}) : ID {acte_id} introuvable, ignoree."
            )
            summary.skipped += 1
            continue

        if existing is not None:
            existing.libelle = libelle
            existing.prix = prix
            existing.code = code or existing.code
            # Categorie fournie => on l'applique ; cellule vide => on conserve
            # la categorie deja en base (meme regle que le code).
            existing.categorie = categorie or existing.categorie
            if not dry_run:
                repo.update_acte(conn, existing)
                # L'etat actif est gere a part de update_acte (cf. set_acte_actif).
                if actif is not None and actif != existing.actif:
                    repo.set_acte_actif(conn, existing.id, actif)
            summary.updated += 1
        else:
            if not dry_run:
                created = repo.create_acte(
                    conn, repo.Acte(id=None, libelle=libelle, prix=prix,
                                    code=code, categorie=categorie)
                )
                if actif is False:  # un acte est cree actif par defaut
                    repo.set_acte_actif(conn, created.id, False)
            summary.created += 1
    return summary


def import_actes(
    path: Path, feuille: Optional[str] = None, dry_run: bool = False, conn=None
) -> ImportSummary:
    """Importe les actes du classeur. N'ecrit rien si dry_run=True.

    `conn` : connexion a reutiliser (ex. celle du serveur). Si None, une connexion
    dediee est ouverte puis fermee (usage CLI).
    """
    rows = _read_rows(path, feuille)
    own = conn is None
    if own:
        conn = connect()
    try:
        return _import_rows(conn, rows, dry_run)
    finally:
        if own:
            conn.close()


# --- Export du referentiel -----------------------------------------------------

# En-tetes du classeur d'export (l'ordre fait foi pour write_template aussi).
_EXPORT_HEADERS = ["ID", "Libelle", "Prix", "Code", "Categorie", "Actif"]


def _build_workbook(actes: list[repo.Acte]):
    """Construit le classeur d'export a partir d'une liste d'actes (deja triee)."""
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Actes"
    ws.append(_EXPORT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for a in actes:
        ws.append([
            a.id,
            a.libelle,
            float(a.prix or 0.0),
            a.code or "",
            a.categorie or "",
            "oui" if a.actif else "non",
        ])
    # Largeurs : ID etroit, libelle large, categorie moyenne.
    for col, width in zip("ABCDEF", (8, 36, 12, 14, 20, 8)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"  # garde l'entete visible au defilement
    # La colonne ID ne doit pas etre editee (cle de rapprochement) : on la grise.
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(color="999999")
    return wb


def export_actes(path: Path, include_inactive: bool = False, conn=None) -> int:
    """Exporte le referentiel d'actes vers `path`. Renvoie le nombre d'actes ecrits.

    `include_inactive=False` -> seuls les actes actifs (ce que l'utilisateur voit) ;
    True -> tous, avec la colonne Actif renseignee pour un aller-retour fidele.
    """
    own = conn is None
    if own:
        conn = connect()
    try:
        actes = repo.list_actes(conn, actifs_seulement=not include_inactive,
                                limit=None)
        _build_workbook(actes).save(path)
        return len(actes)
    finally:
        if own:
            conn.close()


def export_actes_bytes(include_inactive: bool = False, conn=None) -> tuple[bytes, int]:
    """Comme export_actes mais renvoie (octets du .xlsx, nombre d'actes) sans fichier
    temporaire — pratique pour un telechargement HTTP."""
    own = conn is None
    if own:
        conn = connect()
    try:
        actes = repo.list_actes(conn, actifs_seulement=not include_inactive,
                                limit=None)
        buf = BytesIO()
        _build_workbook(actes).save(buf)
        return buf.getvalue(), len(actes)
    finally:
        if own:
            conn.close()


# --- Generation d'un modele Excel ----------------------------------------------

def write_template(path: Path) -> None:
    """Ecrit un classeur VIDE a remplir : seulement la ligne d'entete (en gras),
    colonnes elargies. Meme forme que l'export (colonne ID en tete, a laisser vide
    pour de nouveaux actes). Libelle obligatoire ; Prix au format francais accepte
    (« 1 800,00 ») ; Code / Categorie / Actif facultatifs (la Categorie classe les
    actes, ex. « Prothese », « Chirurgie » ; Actif = oui/non)."""
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Actes"
    ws.append(_EXPORT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col, width in zip("ABCDEF", (8, 36, 12, 14, 20, 8)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"  # garde l'entete visible au defilement
    wb.save(path)


# --- CLI -----------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="crm.import_actes",
        description="Importe / exporte le referentiel d'actes (id, libelle, prix, code, categorie, actif) en .xlsx.",
    )
    parser.add_argument("fichier", nargs="?", help="Classeur .xlsx a importer.")
    parser.add_argument("--feuille", help="Nom de la feuille (defaut : la premiere).")
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Simulation : affiche ce qui serait fait sans rien ecrire en base.",
    )
    parser.add_argument(
        "--export", metavar="FICHIER.xlsx",
        help="Exporte le referentiel vers ce fichier, puis quitte.",
    )
    parser.add_argument(
        "--inclure-inactifs", action="store_true",
        help="Avec --export : inclut aussi les actes desactives.",
    )
    parser.add_argument(
        "--modele", metavar="FICHIER.xlsx",
        help="Genere un fichier modele vierge a remplir, puis quitte.",
    )
    args = parser.parse_args(argv)

    def _say(level: str, text: str) -> None:
        fn = getattr(ui, level, None) if ui is not None else None
        if fn is not None:
            fn(text)
        else:
            print(text)

    # Mode export.
    if args.export:
        try:
            n = export_actes(Path(args.export), include_inactive=args.inclure_inactifs)
        except Exception as exc:  # noqa: BLE001
            _say("error", f"Echec de l'export : {exc}")
            return 1
        _say("success", f"Export termine : {n} acte(s) -> {args.export}")
        return 0

    # Mode generation de modele.
    if args.modele:
        try:
            write_template(Path(args.modele))
        except Exception as exc:  # noqa: BLE001
            _say("error", f"Echec de generation du modele : {exc}")
            return 1
        _say("success", f"Modele cree : {args.modele}")
        return 0

    if not args.fichier:
        parser.error(
            "indiquez un fichier .xlsx a importer (ou --export / --modele)."
        )

    path = Path(args.fichier)
    if not path.is_file():
        _say("error", f"Fichier introuvable : {path}")
        return 1

    if ui is not None:
        ui.banner("Import du referentiel d'actes", str(path))

    # Sauvegarde avant ecriture (jamais en simulation).
    if not args.dry_run:
        dest = backup_db()
        if dest is not None:
            _say("note", f"Sauvegarde prealable : {dest}")
        elif not default_db_path().exists():
            _say("note", "Base inexistante : elle sera creee.")

    try:
        summary = import_actes(path, feuille=args.feuille, dry_run=args.dry_run)
    except RuntimeError as exc:
        _say("error", str(exc))
        return 1

    if args.dry_run:
        _say("warn", "SIMULATION : aucune modification ecrite en base.")

    if summary.errors:
        if ui is not None:
            ui.section("Lignes ignorees")
        for msg in summary.errors:
            _say("warn", msg)

    if ui is not None:
        print()
        ui.success("Import termine." if not args.dry_run else "Simulation terminee.")
        ui.stat("Actes crees", summary.created, accent=ui.GREEN)
        ui.stat("Actes mis a jour", summary.updated, accent=ui.GREEN)
        ui.stat("Lignes ignorees", summary.skipped,
                accent=ui.YELLOW if summary.skipped else "")
    else:
        print("Import termine." if not args.dry_run else "Simulation terminee.")
        print(f"  Actes crees : {summary.created}")
        print(f"  Actes mis a jour : {summary.updated}")
        print(f"  Lignes ignorees : {summary.skipped}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
